from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.styles import Alignment
from PIL import Image as PILImage
import io, base64, os, json, smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
CORS(app)

TEMPLATE = os.path.join(os.path.dirname(__file__), 'AVISO_DE_RECEPCION_.xlsx')

# Correo — configurar via variables de entorno en Render
SMTP_HOST     = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT     = int(os.environ.get('SMTP_PORT', 587))
SMTP_USER     = os.environ.get('SMTP_USER', '')   # tu correo Gmail
SMTP_PASS     = os.environ.get('SMTP_PASS', '')   # contraseña de aplicación Gmail
CORREO_DEST   = 'kurtcortes20154@gmail.com'

def enviar_aviso(rec):
    """Envía correo de aviso de recepción. Falla silenciosamente si no hay credenciales."""
    if not SMTP_USER or not SMTP_PASS:
        return
    try:
        tipo   = rec.get('tipo', '')
        modelo = rec.get('modelo', '')
        marca  = rec.get('marca', '')
        codigo = rec.get('codigo', '')
        cliente= rec.get('cliente', '')
        guia   = rec.get('guia', '')
        receptor = rec.get('receptor', '')

        componente_str = ' '.join(filter(None, [tipo, modelo, marca]))

        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'AVISO DE RECEPCIÓN'
        msg['From']    = SMTP_USER
        msg['To']      = CORREO_DEST

        texto = (
            f"AVISO DE RECEPCIÓN — Swanson Industries\n\n"
            f"Se han recepcionado los siguientes componentes:\n\n"
            f"{componente_str} - {codigo}\n\n"
            f"Cliente:          {cliente}\n"
            f"Guía despacho:    {guia}\n"
            f"Recepcionado por: {receptor}\n\n"
            f"Ingrese al sistema para asignar la OT correspondiente."
        )
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
          <div style="background:#1F3864;padding:20px;text-align:center">
            <h2 style="color:#fff;margin:0">AVISO DE RECEPCIÓN</h2>
            <p style="color:#a8c4e0;margin:4px 0 0">Swanson Industries</p>
          </div>
          <div style="padding:24px;background:#f4f5f7">
            <p style="font-size:15px;color:#444">Se han recepcionado los siguientes componentes:</p>
            <div style="background:#fff;border-left:4px solid #1F3864;padding:16px;margin:16px 0;border-radius:4px">
              <p style="font-size:16px;font-weight:bold;color:#1F3864;margin:0">{componente_str}</p>
              <p style="font-family:monospace;font-size:14px;color:#2E75B6;margin:6px 0 0">{codigo}</p>
            </div>
            <table style="width:100%;font-size:13px;color:#555">
              <tr><td style="padding:4px 0"><b>Cliente:</b></td><td>{cliente}</td></tr>
              <tr><td style="padding:4px 0"><b>Guía despacho:</b></td><td>{guia}</td></tr>
              <tr><td style="padding:4px 0"><b>Recepcionado por:</b></td><td>{receptor}</td></tr>
            </table>
            <div style="margin-top:20px;padding:12px;background:#E8EDF5;border-radius:4px;font-size:12px;color:#666">
              Ingrese al sistema de Recepción de Componentes para asignar la OT correspondiente.
            </div>
          </div>
        </div>"""

        msg.attach(MIMEText(texto, 'plain'))
        msg.attach(MIMEText(html, 'html'))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, CORREO_DEST, msg.as_string())
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")  # log pero no interrumpe el flujo

# ── FOTO SLOTS ────────────────────────────────────────────────────────────
PHOTO_SLOTS_REG = [
    ('recepcion',        2,  54, 10, 62),   # C55:J62
    ('anclaje_vastago',  12, 54, 20, 62),   # M55:T62
    ('anclaje_botella',  2,  63, 10, 71),   # C64:J71
    ('alimentaciones',   12, 63, 20, 71),   # M64:T71
    ('color_componente', 2,  72, 10, 80),   # C73:J80
    ('id_cliente',       12, 72, 20, 80),   # M73:T80
    ('danos_visibles',   2,  81, 10, 89),   # C82:J89
]

def _insert_image(ws, img_data, from_col, from_row, to_col, to_row):
    if isinstance(img_data, str):
        if ',' in img_data: img_data = img_data.split(',')[1]
        img_bytes = base64.b64decode(img_data)
    else:
        img_bytes = img_data
    pil = PILImage.open(io.BytesIO(img_bytes)).convert('RGB')
    buf = io.BytesIO()
    pil.save(buf, format='JPEG', quality=85)
    buf.seek(0)
    img = XLImage(buf)
    img.anchor = TwoCellAnchor(
        editAs='twoCell',
        _from=AnchorMarker(col=from_col, row=from_row, colOff=0, rowOff=0),
        to=AnchorMarker(col=to_col, row=to_row, colOff=0, rowOff=0)
    )
    ws.add_image(img)

def build_fsgi249(rec, photos):
    wb = load_workbook(TEMPLATE)
    ws = wb.active

    def w(addr, val):
        cell = ws[addr]
        if type(cell).__name__ == 'MergedCell': return
        cell.value = val

    ws['U2'].value = 'FSGI-249'
    ws['U3'].value = f"Revisión: 3  ·  {rec.get('codigo','')}"
    ws['U4'].value = f"Fecha: {rec.get('fecha','')}"
    w('D7',  rec.get('cliente',''))
    w('Q7',  f"{rec.get('fecha','')}  {rec.get('hora','')}")
    w('D8',  rec.get('oc','') or '')
    w('Q8',  rec.get('pat_cam','') or '')
    w('F9',  rec.get('guia',''))
    w('Q9',  rec.get('pat_rem','') or '')
    w('E10', rec.get('ot_ant','') or '')
    w('O10', rec.get('ot','') or 'Pendiente')
    ws['B11'].value = (
        f"COMPONENTE:  {rec.get('tipo','')}   Marca: {rec.get('marca','—')}   "
        f"Modelo: {rec.get('modelo','—')}   S/N: {rec.get('nserie','—')}   "
        f"P/N: {rec.get('nparte','—')}   Estado: {rec.get('estvis','—')}   "
        f"Embalaje: {rec.get('cemb','—')}"
    )
    conf = rec.get('estvis','') == 'Bueno'
    ws['G12'].value = 'CONFORME:' + ('   ✓' if conf else '')
    ws['M12'].value = 'NO CONFORME:' + ('   ✓' if not conf else '')

    ck = rec.get('ck', {})
    for key, row in {'11':16,'12':17,'13':18,'21':20,'22':21,'23':22,'31':23}.items():
        val = ck.get(key)
        w(f'M{row}', '✓' if val == 'si' else '')
        w(f'N{row}', '✓' if val == 'no' else '')
        oc = ws[f'O{row}']
        if type(oc).__name__ != 'MergedCell': oc.value = ck.get(f'obs_{key}','')

    # Ítem 4 — añadir el color del componente al costado del detalle de zuncho
    color = (rec.get('color') or '').strip()
    if color:
        base24 = str(ws['C24'].value or 'TIPO DE ELEMENTO: ZUNCHO METALICO').rstrip()
        ws['C24'].value = f"{base24}     Color: {color}"

    if photos.get('guia_recepcion'):
        _insert_image(ws, photos['guia_recepcion'], 1, 26, 21, 52)
    else:
        lineas = [f"Transportista: {rec.get('transp','—')}     Bultos: {rec.get('bultos','1')}     Código: {rec.get('codigo','')}"]
        if rec.get('acc'):      lineas.append(f"Accesorios: {rec['acc']}")
        if rec.get('obs_gral'): lineas.append(f"Obs. generales: {rec['obs_gral']}")
        if rec.get('obs_comp'): lineas.append(f"Obs. componente: {rec['obs_comp']}")
        ws['B27'].value = '\n'.join(lineas)
        ws['B27'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

    for key, fc, fr, tc, tr in PHOTO_SLOTS_REG:
        if photos.get(key):
            _insert_image(ws, photos[key], fc, fr, tc, tr)

    w('D206', rec.get('receptor',''))
    w('N206', rec.get('cargo','Bodega'))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

@app.route('/generar-fsgi249', methods=['POST'])
def generar():
    try:
        rec = json.loads(request.form.get('rec','{}'))
        photos = {}
        for k in ['recepcion','anclaje_vastago','anclaje_botella','alimentaciones','guia_recepcion','color_componente','id_cliente','danos_visibles']:
            f = request.files.get(k)
            if f: photos[k] = f.read()
        xlsx = build_fsgi249(rec, photos)
        # Enviar correo en background (no bloquea la respuesta)
        enviar_aviso(rec)
        codigo  = rec.get('codigo','RC')
        cliente = rec.get('cliente','').replace(' ','_')
        return send_file(xlsx,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{codigo}_{cliente}_FSGI249.xlsx")
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
